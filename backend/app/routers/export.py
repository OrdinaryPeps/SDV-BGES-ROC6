from fastapi import APIRouter, HTTPException, Depends, Response
from typing import Optional
import csv
import io
from datetime import datetime, timezone
from ..core.database import get_db
from ..core.deps import get_current_user
from ..models.user import User
from .performance import filter_tickets

router = APIRouter()

@router.get("/tickets")
async def export_tickets(
    format: str = "csv",
    current_user: User = Depends(get_current_user),
    db = Depends(get_db)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
        
    tickets = await db.tickets.find({}, {"_id": 0}).to_list(10000)
    
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Ticket Number", "Status", "Category", "Description", 
            "Assigned Agent", "Created At", "Completed At", "User Telegram"
        ])
        
        for t in tickets:
            writer.writerow([
                t.get('ticket_number'),
                t.get('status'),
                t.get('category'),
                t.get('description'),
                t.get('assigned_agent_name', 'Unassigned'),
                t.get('created_at'),
                t.get('completed_at', ''),
                t.get('user_telegram_name', '')
            ])
            
        return Response(content=output.getvalue(), media_type="text/csv")
    
    return Response(content="Format not supported", status_code=400)

@router.get("/performance")
async def export_performance(
    year: Optional[str] = None,
    month: Optional[str] = None,
    category: Optional[str] = None,
    agent_id: Optional[str] = None,
    format: str = "csv",
    current_user: User = Depends(get_current_user),
    db = Depends(get_db)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
        
    tickets = await db.tickets.find({}).to_list(10000)
    filtered_tickets = filter_tickets(tickets, year, month, category, agent_id)
    
    # Fetch all comments for the filtered tickets to map to NOTE
    # Optimization: Fetch all comments for these ticket IDs in one go if possible, 
    # but for simplicity and to avoid complex aggregation queries in this step, we might query per ticket or fetch all.
    # Given 10000 tickets limit, fetching all comments might be heavy. 
    # Let's fetch comments where ticket_id in [list of ids].
    
    ticket_ids = [str(t.get('_id')) for t in filtered_tickets]
    comments_cursor = db.comments.find({"ticket_id": {"$in": ticket_ids}})
    comments_map = {}
    async for comment in comments_cursor:
        tid = comment.get('ticket_id')
        if tid not in comments_map:
            comments_map[tid] = []
        comments_map[tid].append(comment)

    if format == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed Report"
        
        headers = [
            "TANGGAL OPEN", "TIKET LAPORAN", "PRODUCT", "TIPE TRANSAKSI", "PERMINTAAN", 
            "ORDER", "WONUM", "TIKET FO", "ND INTERNET/VOICE/SID", "PASSWORD", 
            "PAKET INET", "SN LAMA", "SN BARU", "SN AP", "MAC AP", "SSID", 
            "TIPE ONT", "GPON SLOT/PORT", "VLAN", "SVLAN", "CVLAN", "TASK BIMA", 
            "OWNERGROUP", "KETERANGAN LAINNYA", "HD ROC", "ACTION", "NOTE", 
            "TANGGAL UPDATE", "PELAPOR", "ID PELAPOR", "DURASI TIKET", 
            "MENIT TOTAL", "KAT DURASI", "PERIODE"
        ]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = thin_border
            
        from .performance import parse_datetime
        
        for t in filtered_tickets:
            created_at = parse_datetime(t.get('created_at'))
            completed_at = parse_datetime(t.get('completed_at'))
            
            # Format Dates
            tanggal_open = created_at.strftime("%d/%m/%Y %H:%M") if created_at else ""
            tanggal_update = completed_at.strftime("%d/%m/%Y %H:%M") if completed_at else ""
            
            # Duration Calculation
            durasi_tiket = ""
            menit_total = 0.0
            kat_durasi = ""
            
            if created_at and completed_at:
                duration_seconds = (completed_at - created_at).total_seconds()
                hours, remainder = divmod(duration_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                durasi_tiket = f"{int(hours)}:{int(minutes):02}:{int(seconds):02}"
                
                menit_total = round(duration_seconds / 60, 2)
                
                hours_float = duration_seconds / 3600
                if hours_float < 1:
                    kat_durasi = "< 1 JAM"
                elif 1 <= hours_float < 2:
                    kat_durasi = "1-2 JAM"
                elif 2 <= hours_float <= 3:
                    kat_durasi = "2-3 JAM"
                else:
                    kat_durasi = "> 3 JAM"
            
            # Periode (Month-Year)
            periode = created_at.strftime("%b-%y") if created_at else ""
            
            # Comments for NOTE
            t_comments = comments_map.get(str(t.get('_id')), [])
            # Sort by timestamp
            t_comments.sort(key=lambda x: x.get('timestamp', ''))
            note_parts = []
            for c in t_comments:
                role = "Agent" if c.get('role') == 'agent' else "User"
                note_parts.append(f"[{role}]: {c.get('comment')}")
            note = " | ".join(note_parts)

            row_data = [
                tanggal_open,                               # TANGGAL OPEN
                t.get('ticket_number', ''),                 # TIKET LAPORAN
                t.get('category', ''),                      # PRODUCT
                t.get('tipe_transaksi', ''),                # TIPE TRANSAKSI
                t.get('permintaan', ''),                    # PERMINTAAN
                t.get('order_number', ''),                  # ORDER
                t.get('wonum', ''),                         # WONUM
                t.get('tiket_fo', ''),                      # TIKET FO
                t.get('nd_internet_voice', ''),             # ND INTERNET/VOICE/SID
                t.get('password', ''),                      # PASSWORD
                t.get('paket_inet', ''),                    # PAKET INET
                t.get('sn_lama', ''),                       # SN LAMA
                t.get('sn_baru', ''),                       # SN BARU
                t.get('sn_ap', ''),                         # SN AP
                t.get('mac_ap', ''),                        # MAC AP
                t.get('ssid', ''),                          # SSID
                t.get('tipe_ont', ''),                      # TIPE ONT
                t.get('gpon_slot_port', ''),                # GPON SLOT/PORT
                t.get('vlan', ''),                          # VLAN
                t.get('svlan', ''),                         # SVLAN
                t.get('cvlan', ''),                         # CVLAN
                t.get('task_bima', ''),                     # TASK BIMA
                t.get('ownergroup', ''),                    # OWNERGROUP
                t.get('keterangan_lainnya', ''),            # KETERANGAN LAINNYA
                t.get('assigned_agent_name', ''),           # HD ROC
                "DONE" if t.get('status') == 'completed' else t.get('status', ''), # ACTION
                note,                                       # NOTE
                tanggal_update,                             # TANGGAL UPDATE
                t.get('user_telegram_name', ''),            # PELAPOR
                t.get('user_telegram_id', ''),              # ID PELAPOR
                durasi_tiket,                               # DURASI TIKET
                menit_total,                                # MENIT TOTAL
                kat_durasi,                                 # KAT DURASI
                periode                                     # PERIODE
            ]
            ws.append(row_data)
            
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(), 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=performance_report.xlsx"}
        )
    
    return Response(content="Format not supported", status_code=400)
